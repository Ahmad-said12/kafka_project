import asyncio
import json
from aiokafka import AIOKafkaConsumer
from openpyxl import load_workbook, Workbook
import os

EXCEL_FILE = "data3.xlsx"

def save_to_excel(name: str):
    wb = None
    if os.path.exists(EXCEL_FILE) and os.path.getsize(EXCEL_FILE) > 0:
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        except Exception:
            wb = None
            
    if wb is None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Items"
        ws.append(["Name"])  

    ws.append([name])
    wb.save(EXCEL_FILE)
    print(f"✅ save on Excel: {name}")

async def consume():
    consumer = AIOKafkaConsumer(
        "items_topic3",
        bootstrap_servers="localhost:9092",
        group_id="items_group"
    )
    await consumer.start()
    print("✅ Consumer is running and listening to 'items_topic3'...")
    try:
        async for msg in consumer:
            data = json.loads(msg.value.decode("utf-8"))
            await asyncio.to_thread(save_to_excel, data["name"])
    finally:
        await consumer.stop()

if __name__ == "__main__":
    asyncio.run(consume())